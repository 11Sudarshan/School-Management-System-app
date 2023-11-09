import openpyxl
from flask import Flask, request, jsonify
from pymongo import MongoClient
from bson.json_util import dumps
# Open Excel file
workbook = openpyxl.load_workbook('./server/db/stud.xlsx')

client = MongoClient("mongodb+srv://kiafbial123:kiafbial123@cluster0.o3lbxfj.mongodb.net/?retryWrites=true&w=majority")
db = client.ghps_aradeshanahalli

# Select worksheet
worksheet = workbook['8th std']
c=0
# Get dimensions of worksheet
max_row = worksheet.max_row
max_col = worksheet.max_column

# Iterate over each row in the worksheet
for row in range(6, max_row+1):
    # Create empty list to hold row data
    row_data = []
    # Iterate over each cell in the row
    for col in range(1, max_col+1):
        # Get cell value and append to row data list
        cell_value = str(worksheet.cell(row=row, column=col).value)
        # print(cell_value)
        # print(type(cell_value))
        if(cell_value == "None"):
            cell_value = "No Data"
        # if(cell_value!= "No Data"):
        row_data.append(cell_value)

    sl_no = row_data[0]
    name = row_data[1]
    fname = row_data[2]
    mname = row_data[3]
    sats_no = row_data[4]
    habitation =row_data[5]
    aadhar_no = row_data[6]
    addsm_no = row_data[7]
    doa = row_data[8]
    dob = row_data[9]
    gender = row_data[10]
    caste = row_data[11]
    minority =row_data[12]
    acc_no = row_data[12]
    ifsc_code = row_data[14]
    branch = row_data[15]
    bpl_no = row_data[16]
    ph_no = row_data[17]
    bhagya_yojana = row_data[18]
    fd_no = row_data[19]
    addsm_rti =row_data[20]
    Class = row_data[21]
    prv_class =row_data[22]
    medium = row_data[23]
    height = row_data[24]
    weight = row_data[25]
    med_advice = row_data[26]
    place = 'No Data'
    # print(row_data)
    if(name!= "No Data"):
        c+=1
        result = db.students.insert_one({'sl_no': sl_no, 'name': name, 'fname': fname, 'mname': mname, 'sats_no': sats_no, 'habitation': habitation, 'aadhar_no': aadhar_no, 'addsm_no': addsm_no, 'doa': doa, 'dob': dob, 'gender': gender, 'caste': caste, 'minority': minority, 'acc_no': acc_no, 'ifsc_code': ifsc_code, 'branch': branch, 'bpl_no': bpl_no, 'ph_no': ph_no, 'bhagya_yojana': bhagya_yojana, 'fd_no': fd_no, 'addsm_rti': addsm_rti, 'class_': Class, 'prv_class': prv_class, 'medium': medium, 'height': height, 'weight': weight, 'med_advice': med_advice,'place_name':place})
    

